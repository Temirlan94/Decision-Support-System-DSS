from tkinter import *
from tkinter import messagebox
from PIL import Image, ImageTk
import datetime as dt
import pandas as pd
import numpy as np
import docx2txt
import os
import pyowm
import urllib.request
import socket
import platform

import speech_recognition as SR
import pyttsx3
import pywhatkit
import wikipedia
import openpyxl
from openpyxl import load_workbook
virtHelp = pyttsx3.init()


class Create_Window:
# Main menu intitialization===========================================================================

    def __init__(self, window):

        self.window = window
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        self.window.geometry(f'{screen_width}x{screen_height}')
        self.window.title("DSS_MENU")
        self.window.configure(bg="#049fd9")
        w=int(screen_width)
        h=int(screen_height)

        try:
            self.bgpic = ImageTk.PhotoImage(Image.open("pics/bgg.jpg").resize((w, h), Image.ANTIALIAS))
            self.background_label = Label(image=self.bgpic)
            self.background_label.pack()
        except FileNotFoundError:
            self.background_label=Label(bg="#049fd9")
            self.background_label.pack()

        try:
            self.susmed_logo = ImageTk.PhotoImage(Image.open("pics/susmedhouse_logo.png").resize((int(w/4.25), int(h/5.70)), Image.ANTIALIAS))
            self.susmed_label = Label(image=self.susmed_logo)
            self.susmed_label.place(x=w/1.35, y=h/1.35)
        except FileNotFoundError:
            pass

        try:
            self.artecs_logo = ImageTk.PhotoImage(Image.open("pics/artecs_logo.png").resize((int(w/4.25), int(h/5.70)), Image.ANTIALIAS))
            self.artecs_label = Label(image=self.artecs_logo)
            self.artecs_label.place(x=w/1.35, y=h/1.80)
        except FileNotFoundError:
            pass

        # Date&Time intitialization===========================================================================
        self.date = Label(window, text=f"{dt.datetime.now():%a, %b %d %Y}", fg="white", bg="#049fd9", font=("helvetica", int(w/64)))
        self.date.place(x=w/1.25, y=h/16.5)

        self.currentTime=f"{dt.datetime.now():%H:%M %p}"
        self.time = Label(window, text=self.currentTime, fg="white", bg="#049fd9", font=("helvetica",int(w/64)))
        self.time.place(x=w/1.25, y=h/9.5)


        # Button creation=====================================================================================
        self.ePrice = Button(window, text="Electricity Prices", fg="blue", bg="#fceea7",width=int(w/102.4), height=int(h/288), font=int(w / 102), command=self.ePriceShow)
        self.ePrice.place(x=w/50, y=h/50)

        self.cropPrices = Button(window, text="Crop Prices", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288),font=int(w/102),command=self.cropPriceShow)
        self.cropPrices.place(x=w/50, y=h/10)
        self.cropWindow=None #not to open same window multiple times

        self.cropPrices = Button(window, text="Greenhouse conditions", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288), font=int(w/102),command=self.greenConditionShow)
        self.cropPrices.place(x=w/50, y=h/5.5)

        self.cropPrices = Button(window, text="Crop status", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288),font=int(w/102),command=self.cropStatus)
        self.cropPrices.place(x=w/50, y=h/3.8)

        self.cropPrices = Button(window, text="Greenhouse controls", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288), font=int(w/102),command=self.greenHouseControls)
        self.cropPrices.place(x=w/5, y=h/50)

        self.cropPrices = Button(window, text="Feasibility report", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288), font=int(w/102),command=self.feasibilityReportShow)
        self.cropPrices.place(x=w/5, y=h/10)

        self.cropPrices = Button(window, text="Instructions", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288),font=int(w/102),command=self.instructionShow)
        self.cropPrices.place(x=w/5, y=h/5.5)

        self.cropPrices = Button(window, text="Warnings", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288), font=int(w/102),command=self.greenHouseWarnings)
        self.cropPrices.place(x=w/5, y=h/3.8)

        self.sysInfo = Button(window, text="Device Info", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288), font=int(w/102), command=self.sysInfo)
        self.sysInfo.place(x=w/50, y=h/2.8)

        self.appInfo = Button(window, text="Application Info", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288), font=int(w/102),command=self.appInfo)
        self.appInfo.place(x=w/5, y=h/2.8)

        self.virtualAssistant = Button(window, text="Virtual Assistant (Beta)", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288), font=int(w/102),command=self.virtualAssistant)
        self.virtualAssistant.place(x=w/50, y=h/2.2)

        self.close_button = Button(window, text="Close", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288),font=int(w/102), command=window.quit)
        self.close_button.place(x=w/50, y=h/1.4)

################################################
# System info button
################################################

    def sysInfo(self):
        try:
            messagebox.showinfo("Device Info", "Platform: {} \nPlatform Release: {}\nPlatform Version: {}\nArchitecture: {}"
            "\nHostname: {}\nIPv4 address: {}\nProcessor: {}".format(platform.system(),platform.release(),platform.version(),platform.machine(),socket.gethostname(),socket.gethostbyname(socket.gethostname()),platform.processor()))
        except:
            messagebox.showinfo("Device Info","Uknown")

################################################
# Application info button
################################################

    def appInfo(self):
        messagebox.showinfo("Application Info","Name: SusmedHouse DSS \nVersion: 1.0\nPrepared by: AR&TeCS\nDeveloper: Temirlan Bayeshov\nContact E-Mail: info@ar-tecs.com\nContact Phone: +90 (312) 484 55 15\nCountry:Turkey")

################################################
# Virtual Assistant button
################################################

    def virtualAssistant(self):
        def Virt_speak(content):
            virtHelp.say(content)
            virtHelp.runAndWait()
            print(content)

        listener = SR.Recognizer()

        def listen_to_user():
            try:
                Virt_speak("Hey there! I'm Susmedhouse, your virtual assistant.")
                with SR.Microphone() as source:
                    Virt_speak("How can I help you?")
                    user_audio = listener.listen(source)
                    user_input = listener.recognize_google(user_audio).lower()
                    if "james" in user_input:
                        print(user_input.upper())
                        user_input = user_input.replace("james", "")
            except:
                pass
            return user_input

        command = listen_to_user()
        if "play" in command:
            command = command.replace("play", "")
            Virt_speak("Playing " + command)
            pywhatkit.playonyt(command)

        elif "weather" in command:
            APIKEY = 'ed4a6d93e2c3f1bbbcd7b31833885c0b'  # your API Key here as string
            OpenWMap = pyowm.OWM(APIKEY)  # Use API key to get data
            Weather = OpenWMap.weather_at_place('Ankara')  # give where you need to see the weather
            Data = Weather.get_weather()
            temp = Data.get_temperature(unit='celsius')
            humidity = Data.get_humidity()
            wind = Data.get_wind()
            cloud = Data.get_clouds()

            Virt_speak("Here is the weather outside:")
            Virt_speak("Temperature is {}".format(str(temp['temp']))+"Degrees")
            Virt_speak("Humidity is {}".format(str(humidity))+"Percent")
            Virt_speak("Wind speed is {}".format(str(wind['speed']))+ "meter per second")
        elif "company" in command:
            Virt_speak("ARTECS Anadolu AR-GE Technology Engineering and Consulting Company. To contribute to the development of advanced technological products in 2013, to make R&D studies, to create new employment areas by "
                       "utilizing the developed technologies, to develop joint projects with foreign research centers specialized in advanced technology "
                       "and or to interface with the related industrial organizations with the aim of creating")
        elif "project" in command:
            Virt_speak("Sustainability and Competitiveness of Mediterranean Greenhouse and Intensive Horticulture.This project is part of the PRIMA programme supported by the European Union")

        elif "who am i" in command:
            Virt_speak("Im SusmedHouse Virtual Assistant, I will try to help and answer your questions! I was developed in 2021 in Ankara, Turkey by Temirlan Bayeshov")

        elif "thank you" in command:
            Virt_speak("Thank you too! Bye Bye!")

        elif "bye" in command:
            Virt_speak("Thank you ! Bye Bye!")

        elif "how are you" in command:
            Virt_speak("Thank you for asking. Im excellent today! Wish you a good and productive day!")

        elif "help" in command:
            Virt_speak("You can use keywords like")
            Virt_speak("Weather")
            Virt_speak("Company")
            Virt_speak("Project")
            Virt_speak("who am i")
            Virt_speak("Help")
            Virt_speak("For futher information conntact developer")

        else:
            Virt_speak("Searching for" + command)
            info = wikipedia.summary(command, 1)
            Virt_speak(info)

################################################
# Electricity Price Button
# Local information
# Will be taken from API
################################################
    def ePriceShow(self):
            currentMonth = text = f"{dt.datetime.now():%b}"
            currentYear= int(f"{dt.datetime.now():%Y}")
            try:
                ePrice_data = pd.read_excel(r'excelFiles/elektrik.xls')
                all_data = pd.DataFrame(ePrice_data)
                all_data['Condition1'] = all_data['Month'].apply(lambda x: 'True' if x == currentMonth else 'False')
                all_data['Condition2'] = all_data['Year'].apply(lambda x: 'True' if x == currentYear else 'False')
                all_data['Condition3'] = np.where((all_data['Condition1'] == all_data['Condition2']), all_data['Condition1'], "False")
                showPrice = all_data.loc[all_data['Condition3'] == 'True']['Price'].values[0]
                messagebox.showinfo("showinfo", "{} {} Price is: {} TL".format(currentMonth,currentYear,showPrice))
                #print (all_data)
            except FileNotFoundError:
                messagebox.showinfo("showinfo", "No prices available, try again later!")
                datetime_object = str(dt.datetime.now())
                strIssue = "Electricity price not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass

################################################
# Crop Price Button
# Local information
# Will be taken from API
################################################

    def cropPriceShow(self):

        if self.cropWindow is None:  #not to open same window multiple times
            currentMonth = text = f"{dt.datetime.now():%b}"
            currentYear= int(f"{dt.datetime.now():%Y}")

            def selection1():
                try:
                    tomatoPrice = pd.read_excel(r'excelFiles/tomato.xls')
                    all_data = pd.DataFrame(tomatoPrice)
                    all_data['Condition1'] = all_data['Month'].apply(lambda x: 'True' if x == currentMonth else 'False')
                    all_data['Condition2'] = all_data['Year'].apply(lambda x: 'True' if x == currentYear else 'False')
                    all_data['Condition3'] = np.where((all_data['Condition1'] == all_data['Condition2']),all_data['Condition1'], "False")
                    showPrice = all_data.loc[all_data['Condition3'] == 'True']['Price'].values[0]
                    messagebox.showinfo("showinfo", "{} {} Tomato price is: {} TL".format(currentMonth,currentYear,showPrice))
                except FileNotFoundError:
                    messagebox.showinfo("showinfo", "No prices available, try again later!")
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Tomato price not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass

                self.cropWindow.destroy()
                self.cropWindow = None
            def selection2():
                try:
                    lettucePrice = pd.read_excel(r'excelFiles/lettuce.xls')
                    all_data = pd.DataFrame(lettucePrice)
                    all_data['Condition1'] = all_data['Month'].apply(lambda x: 'True' if x == currentMonth else 'False')
                    all_data['Condition2'] = all_data['Year'].apply(lambda x: 'True' if x == currentYear else 'False')
                    all_data['Condition3'] = np.where((all_data['Condition1'] == all_data['Condition2']),all_data['Condition1'], "False")
                    showPrice = all_data.loc[all_data['Condition3'] == 'True']['Price'].values[0]
                    messagebox.showinfo("showinfo", "{} {} Lettuce price is: {} TL".format(currentMonth,currentYear,showPrice))
                except FileNotFoundError:
                    messagebox.showinfo("showinfo", "No prices available, try again later!")
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Lettuce price not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass

                self.cropWindow.destroy()
                self.cropWindow = None
            def selection3():
                try:
                    pepperPrice = pd.read_excel(r'excelFiles/pepper.xls')
                    all_data = pd.DataFrame(pepperPrice)
                    all_data['Condition1'] = all_data['Month'].apply(lambda x: 'True' if x == currentMonth else 'False')
                    all_data['Condition2'] = all_data['Year'].apply(lambda x: 'True' if x == currentYear else 'False')
                    all_data['Condition3'] = np.where((all_data['Condition1'] == all_data['Condition2']),all_data['Condition1'], "False")
                    showPrice = all_data.loc[all_data['Condition3'] == 'True']['Price'].values[0]
                    messagebox.showinfo("showinfo", "{} {} Pepper price is: {} TL".format(currentMonth,currentYear,showPrice))
                except FileNotFoundError:
                    messagebox.showinfo("showinfo", "No prices available, try again later!")
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Pepper price not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass
                self.cropWindow.destroy()
                self.cropWindow = None

            screen_width = self.window.winfo_screenwidth()
            screen_height = self.window.winfo_screenheight()
            w = int(screen_width)
            h = int(screen_height)
            self.cropWindow = Tk()
            self.cropWindow.geometry(f'{int(w / 4.2)}x{int(h / 3.2)}')
            self.cropWindow.configure(bg="#049fd9")
            self.cropWindow.title("Crop Price")


            radio = IntVar()
            R1 = Radiobutton(self.cropWindow, text="Tomato",bg="#fceea7", width=int(w/102.4), height=int(h/288),font=int(w/102), variable=radio, value= 1,command=selection1)
            R1.pack(anchor=W)

            R2 = Radiobutton(self.cropWindow, text="Lettuce",bg="#fceea7", width=int(w/102.4), height=int(h/288),font=int(w/102), variable=radio, value= 2,command=selection2)
            R2.pack(anchor=W)

            R3 = Radiobutton(self.cropWindow, text="Pepper", bg="#fceea7", width=int(w/102.4), height=int(h/288),font=int(w/102), variable=radio, value= 3,command=selection3)
            R3.pack(anchor=W)

################################################
# Instruction button
# Local information from text file
 ################################################
    def instructionShow(self):
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        w = int(screen_width)
        h = int(screen_height)
        try:
            self.instructionBox = Tk()
            self.instructionBox.geometry(f'{int(w / 2.56)}x{int(h / 1.4)}')
            labelframe1 = LabelFrame(self.instructionBox, text="Instructions")
            labelframe1.pack(fill="both", expand="yes")
            result = docx2txt.process("instructions/cleaning.docx")
            #mssg="Clean************************************************"
            toplabel = Label(labelframe1, text=result)
            toplabel.place(x=0,y=50)
        except FileNotFoundError:
            messagebox.showinfo("showinfo", "No instruction available, try again later!")
            datetime_object = str(dt.datetime.now())
            strIssue = "Instruction not found!"
            df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
            try:
                path = 'excelFiles/historyLogs.xlsx'
                book = load_workbook(path)
                writer = pd.ExcelWriter(path, engine='openpyxl')
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                for sheetname in writer.sheets:
                    df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                index=False, header=False)
                writer.save()
            except FileNotFoundError:
                pass

################################################
# Feasibility Report button
# Locad data from word file
################################################
    def feasibilityReportShow(self):
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        w = int(screen_width)
        h = int(screen_height)
        try:
            self.messageBox = Tk()
            self.messageBox.geometry(f'{int(w / 2.56)}x{int(h / 1.4)}')
            ePrice_data = pd.read_excel(r'excelFiles/elektrik.xls')
            all_data = pd.DataFrame(ePrice_data)

            labelframe1 = LabelFrame(self.messageBox, text="Electricity Price")
            labelframe1.pack(fill="both", expand="yes")

            toplabel = Label(labelframe1, text=all_data)
            toplabel.place(x=0, y=h/28.8)
        except FileNotFoundError:
            messagebox.showinfo("showinfo", "No report available, try again later!")

################################################
# Greenhouse conditions button
# Connected to Labview for graph retrieval of rooms
################################################
    def greenConditionShow(self):

        def showImg(room_id):
            if room_id==1:
                self.labelframe = LabelFrame(self.conditionWindow, text="Cabin 1 Status", width=int(w / 1.5),height=int(h))
                self.labelframe.place(x=w / 2.6, y=h / 25)
                try:
                    self.pic1 = ImageTk.PhotoImage(Image.open("PSO_Check/Oda 1_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                    self.pic2 = ImageTk.PhotoImage(Image.open("PSO_Check/Oda 1_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))

                    toplabel = Label(self.labelframe, image=self.pic1)
                    toplabel.place(x=w/70,y=h/60)

                    toplabel1 = Label(self.labelframe, image=self.pic2)
                    toplabel1.place(x=w / 70, y=h / 2.5)

                except FileNotFoundError:
                    toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)",font=int(w/102))
                    toplabel.place(x=w/3.5,y=h/2.5)
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Room1 Graphs not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass


            elif room_id==2:
                self.labelframe = LabelFrame(self.conditionWindow, text="Cabin 2 Status", width=int(w / 1.5),height=int(h))
                self.labelframe.place(x=w / 2.6, y=h / 25)

                try:
                    self.pic1 = ImageTk.PhotoImage(Image.open("PSO_Check/Oda 2_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                    self.pic2 = ImageTk.PhotoImage(Image.open("PSO_Check/Oda 2_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))

                    toplabel = Label(self.labelframe, image=self.pic1)
                    toplabel.place(x=w / 70, y=h / 60)

                    toplabel1 = Label(self.labelframe, image=self.pic2)
                    toplabel1.place(x=w / 70, y=h / 2.5)

                except FileNotFoundError:
                    toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)", font=int(w / 102))
                    toplabel.place(x=w / 3.5, y=h / 2.5)
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Room2 Graphs not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass

            elif room_id == 3:
                self.labelframe = LabelFrame(self.conditionWindow, text="Cabin 3 Status", width=int(w / 1.5),height=int(h))
                self.labelframe.place(x=w / 2.6, y=h / 25)
                try:
                    self.pic1 = ImageTk.PhotoImage(Image.open("PSO_Check/Oda 3_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                    self.pic2 = ImageTk.PhotoImage(Image.open("PSO_Check/Oda 3_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))

                    toplabel = Label(self.labelframe, image=self.pic1)
                    toplabel.place(x=w / 70, y=h / 60)

                    toplabel1 = Label(self.labelframe, image=self.pic2)
                    toplabel1.place(x=w / 70, y=h / 2.5)

                except FileNotFoundError:
                    toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)", font=int(w / 102))
                    toplabel.place(x=w / 3.5, y=h / 2.5)
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Room3 Graphs not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass

            elif room_id == 4:
                self.labelframe = LabelFrame(self.conditionWindow, text="Cabin 4 Status", width=int(w / 1.5),height=int(h))
                self.labelframe.place(x=w / 2.6, y=h / 25)
                try:
                    self.pic1 = ImageTk.PhotoImage(Image.open("PSO_Check/Oda 4_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                    self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 4_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))

                    toplabel = Label(self.labelframe, image=self.pic1)
                    toplabel.place(x=w / 70, y=h / 60)

                    toplabel1 = Label(self.labelframe, image=self.pic2)
                    toplabel1.place(x=w / 70, y=h / 2.5)

                except FileNotFoundError:
                    toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)", font=int(w / 102))
                    toplabel.place(x=w / 3.5, y=h / 2.5)
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Room4 Graphs not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass

            elif room_id == 5:
                self.labelframe = LabelFrame(self.conditionWindow, text="Cabin 5 Status", width=int(w / 1.5), height=int(h))
                self.labelframe.place(x=w / 2.6, y=h / 25)
                try:
                    self.pic1 = ImageTk.PhotoImage(Image.open("PSO_Check/Oda 5_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                    self.pic2 = ImageTk.PhotoImage(Image.open("PSO_Check/Oda 5_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))

                    toplabel = Label(self.labelframe, image=self.pic1)
                    toplabel.place(x=w / 70, y=h / 60)

                    toplabel1 = Label(self.labelframe, image=self.pic2)
                    toplabel1.place(x=w / 70, y=h / 2.5)

                except FileNotFoundError:
                    toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)", font=int(w / 102))
                    toplabel.place(x=w / 3.5, y=h / 2.5)

            elif room_id == 6:
                self.labelframe = LabelFrame(self.conditionWindow, text="Cabin 6 Status", width=int(w / 1.5), height=int(h))
                self.labelframe.place(x=w / 2.6, y=h / 25)
                try:
                    self.pic1 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 6_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                    self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 6_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))

                    toplabel = Label(self.labelframe, image=self.pic1)
                    toplabel.place(x=w / 70, y=h / 60)

                    toplabel1 = Label(self.labelframe, image=self.pic2)
                    toplabel1.place(x=w / 70, y=h / 2.5)

                except FileNotFoundError:
                    toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)", font=int(w / 102))
                    toplabel.place(x=w / 3.5, y=h / 2.5)
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Room6 Graphs not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass

            elif room_id == 7:
                self.labelframe = LabelFrame(self.conditionWindow, text="Cabin 7 Status", width=int(w / 1.5),height=int(h))
                self.labelframe.place(x=w / 2.6, y=h / 25)
                try:
                    self.pic1 = ImageTk.PhotoImage(
                        Image.open("PSO_Check/Oda 7_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                    self.pic2 = ImageTk.PhotoImage(
                        Image.open("PSO_Check/Oda 7_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))

                    toplabel = Label(self.labelframe, image=self.pic1)
                    toplabel.place(x=w / 70, y=h / 60)

                    toplabel1 = Label(self.labelframe, image=self.pic2)
                    toplabel1.place(x=w / 70, y=h / 2.5)

                except FileNotFoundError:
                    toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)", font=int(w / 102))
                    toplabel.place(x=w / 3.5, y=h / 2.5)
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Room7 Graphs not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass

            elif room_id == 8:
                self.labelframe = LabelFrame(self.conditionWindow, text="Cabin 8 Status", width=int(w / 1.5),height=int(h))
                self.labelframe.place(x=w / 2.6, y=h / 25)
                try:
                    self.pic1 = ImageTk.PhotoImage(
                        Image.open("PSO_Check/Oda 8_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                    self.pic2 = ImageTk.PhotoImage(
                        Image.open("PSO_Check/Oda 8_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))

                    toplabel = Label(self.labelframe, image=self.pic1)
                    toplabel.place(x=w / 70, y=h / 60)

                    toplabel1 = Label(self.labelframe, image=self.pic2)
                    toplabel1.place(x=w / 70, y=h / 2.5)


                except FileNotFoundError:
                    toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)", font=int(w / 102))
                    toplabel.place(x=w / 3.5, y=h / 2.5)
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Room8 Graphs not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass

            elif room_id == 9:
                self.labelframe = LabelFrame(self.conditionWindow, text="Cabin 9 Status", width=int(w / 1.5), height=int(h))
                self.labelframe.place(x=w / 2.6, y=h / 25)
                try:
                    self.pic1 = ImageTk.PhotoImage(
                        Image.open("PSO_Check/Oda 9_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                    self.pic2 = ImageTk.PhotoImage(
                        Image.open("PSO_Check/Oda 9_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))

                    toplabel = Label(self.labelframe, image=self.pic1)
                    toplabel.place(x=w / 70, y=h / 60)

                    toplabel1 = Label(self.labelframe, image=self.pic2)
                    toplabel1.place(x=w / 70, y=h / 2.5)

                except FileNotFoundError:
                    toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)", font=int(w / 102))
                    toplabel.place(x=w / 3.5, y=h / 2.5)
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Room9 Graphs not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass

            elif room_id == 10:
                self.labelframe = LabelFrame(self.conditionWindow, text="Cabin 10 Status", width=int(w / 1.5),height=int(h))
                self.labelframe.place(x=w / 2.6, y=h / 25)
                try:
                    self.pic1 = ImageTk.PhotoImage(
                        Image.open("PSO_Check/Oda 10_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                    self.pic2 = ImageTk.PhotoImage(
                        Image.open("PSO_Check/Oda 10_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))

                    toplabel = Label(self.labelframe, image=self.pic1)
                    toplabel.place(x=w / 70, y=h / 60)

                    toplabel1 = Label(self.labelframe, image=self.pic2)
                    toplabel1.place(x=w / 70, y=h / 2.5)

                except FileNotFoundError:
                    toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)", font=int(w / 102))
                    toplabel.place(x=w / 3.5, y=h / 2.5)
                    datetime_object = str(dt.datetime.now())
                    strIssue = "Room10 Graphs not found!"
                    df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                    try:
                        path = 'excelFiles/historyLogs.xlsx'
                        book = load_workbook(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                        index=False, header=False)
                        writer.save()
                    except FileNotFoundError:
                        pass




        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        w = int(screen_width)
        h = int(screen_height)
        self.conditionWindow = Toplevel()    #Toplevel
        self.conditionWindow.geometry(f'{int(w)}x{int(h)}')
        self.conditionWindow.configure(bg="#049fd9")
        self.conditionWindow.title("Greenhouse Conditions")



        self.buttonOda = Button(self.conditionWindow, text="Cabin_1", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102),command=lambda: showImg(1))
        self.buttonOda.place(x=w/55, y=h/25)

        self.buttonOda = Button(self.conditionWindow, text="Cabin_2", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102),command=lambda: showImg(2))
        self.buttonOda.place(x=w/55, y=h/8)

        self.buttonOda = Button(self.conditionWindow, text="Cabin_3", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102),command=lambda: showImg(3))
        self.buttonOda.place(x=w/55, y=h /4.8)

        self.buttonOda = Button(self.conditionWindow, text="Cabin_4", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102),command=lambda: showImg(4))
        self.buttonOda.place(x=w/55, y=h/3.4)

        self.buttonOda = Button(self.conditionWindow, text="Cabin_5", fg="blue", bg="#fceea7", width=int(w / 102.4), height=int(h / 288), font=int(w / 102),command=lambda: showImg(5))
        self.buttonOda.place(x=w / 55, y=h / 2.65)

        self.buttonOda = Button(self.conditionWindow, text="Cabin_6", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102),command=lambda: showImg(6))
        self.buttonOda.place(x=w/5, y=h/25)

        self.buttonOda = Button(self.conditionWindow, text="Cabin_7", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102),command=lambda: showImg(7))
        self.buttonOda.place(x=w/5, y=h/8)

        self.buttonOda = Button(self.conditionWindow, text="Cabin_8", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102),command=lambda: showImg(8))
        self.buttonOda.place(x=w/5, y=h /4.8)

        self.buttonOda = Button(self.conditionWindow, text="Cabin_9", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102),command=lambda: showImg(9))
        self.buttonOda.place(x=w / 5, y=h / 3.4)

        self.buttonOda = Button(self.conditionWindow, text="Cabin_10", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102),command=lambda: showImg(10))
        self.buttonOda.place(x=w / 5, y=h / 2.65)


        self.close_button = Button(self.conditionWindow, text="Close", fg="blue", bg="#fceea7", width=int(w/102.4), height=int(h/288),font=int(w/102), command=self.conditionWindow.destroy)
        self.close_button.place(x=w/50, y=h/1.4)

################################################
# Crop Status button
# Will take data from AI module
################################################
    def cropStatus(self):
################################################
# Crop Status ->3 options for pepper/tomato/lettuce
################################################
        def show_status(plant_id):
            if plant_id==1:
                self.labelframe = LabelFrame(self.statusWindow, text="Tomato", width=int(w / 1.2),height=int(h))
                self.labelframe.place(x=w / 4.6, y=h / 25)
            elif plant_id==2:
                self.labelframe = LabelFrame(self.statusWindow, text="Lettuce", width=int(w / 1.2), height=int(h))
                self.labelframe.place(x=w / 4.6, y=h / 25)

                try:
                    self.pic1 = ImageTk.PhotoImage(Image.open("lettuceWeek/samples/orig.jpg").resize((int(w / 8), int(h /6)),Image.ANTIALIAS))
                    self.pic2 = ImageTk.PhotoImage(Image.open("lettuceWeek/samples/fg1.jpg").resize((int(w / 8), int(h / 6)), Image.ANTIALIAS))
                    self.pic3 = ImageTk.PhotoImage(Image.open("lettuceWeek/samples/fg2.jpg").resize((int(w / 8), int(h / 6)), Image.ANTIALIAS))


                    toplabel = Label(self.labelframe, image=self.pic1)
                    toplabel.place(x=w / 70, y=h / 4)

                    toplabel1 = Label(self.labelframe, image=self.pic2)
                    toplabel1.place(x=w / 6, y=h / 4)

                    toplabel1 = Label(self.labelframe, image=self.pic3)
                    toplabel1.place(x=w / 3.1, y=h / 4)

                    self.week1 = ImageTk.PhotoImage(Image.open("lettuceWeek/artecs_lettuce/1Week.jpg").resize((int(w / 10), int(h / 10)),Image.ANTIALIAS))
                    self.week2 = ImageTk.PhotoImage(Image.open("lettuceWeek/artecs_lettuce/2Week.jpg").resize((int(w / 10), int(h / 10)),Image.ANTIALIAS))
                    self.week3 = ImageTk.PhotoImage(Image.open("lettuceWeek/artecs_lettuce/3Week.jpg").resize((int(w / 10), int(h / 10)),Image.ANTIALIAS))
                    self.week4 = ImageTk.PhotoImage(Image.open("lettuceWeek/artecs_lettuce/4Week.jpg").resize((int(w / 10), int(h / 10)),Image.ANTIALIAS))
                    self.week5 = ImageTk.PhotoImage(Image.open("lettuceWeek/artecs_lettuce/5Week.jpg").resize((int(w / 10), int(h / 10)),Image.ANTIALIAS))
                    self.week6 = ImageTk.PhotoImage(Image.open("lettuceWeek/artecs_lettuce/6Week.jpg").resize((int(w / 10), int(h / 10)),Image.ANTIALIAS))
                    self.week7 = ImageTk.PhotoImage(Image.open("lettuceWeek/artecs_lettuce/7Week.jpg").resize((int(w / 10), int(h / 10)),Image.ANTIALIAS))

                    toplabel = Label(self.labelframe, image=self.week1)
                    toplabel.place(x=w / 70, y=h / 60)
                    toplabel = Label(self.labelframe, text="1 Week", font=int(w / 102))
                    toplabel.place(x=w / 70, y=h / 60)

                    toplabel = Label(self.labelframe, image=self.week2)
                    toplabel.place(x=w / 10, y=h / 60)
                    toplabel = Label(self.labelframe, text="2 Week", font=int(w / 102))
                    toplabel.place(x=w / 10, y=h / 60)

                    toplabel = Label(self.labelframe, image=self.week3)
                    toplabel.place(x=w / 5, y=h / 60)
                    toplabel = Label(self.labelframe, text="3 Week", font=int(w / 102))
                    toplabel.place(x=w / 5, y=h / 60)

                    toplabel = Label(self.labelframe, image=self.week4)
                    toplabel.place(x=w / 3.5, y=h / 60)
                    toplabel = Label(self.labelframe, text="4 Week", font=int(w / 102))
                    toplabel.place(x=w / 3.5, y=h / 60)

                    toplabel = Label(self.labelframe, image=self.week5)
                    toplabel.place(x=w / 2.6, y=h / 60)
                    toplabel = Label(self.labelframe, text="5 Week", font=int(w / 102))
                    toplabel.place(x=w / 2.6, y=h / 60)

                    toplabel = Label(self.labelframe, image=self.week6)
                    toplabel.place(x=w / 2.1, y=h / 60)
                    toplabel = Label(self.labelframe, text="6 Week", font=int(w / 102))
                    toplabel.place(x=w / 2.1, y=h / 60)

                    toplabel = Label(self.labelframe, image=self.week7)
                    toplabel.place(x=w / 1.8, y=h / 60)
                    toplabel = Label(self.labelframe, text="7 Week", font=int(w / 102))
                    toplabel.place(x=w / 1.8, y=h / 60)


                except FileNotFoundError:
                    toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)", font=int(w / 102))
                    toplabel.place(x=w / 3.5, y=h / 2.5)


                try:
                    lsize = pd.read_excel(r'excelFiles/size.xls')
                    all_data = pd.DataFrame(lsize)
                    lettuceW=lsize["Width"].head(1)
                    lettuceH=lsize["Height"].head(1)
                    lettuceA=int(lettuceW*lettuceH)
                    sizeValue=0
                    if lettuceA>50 and lettuceA <=54:
                        toplabel = Label(self.labelframe, text="1 Week", font=int(w / 102))
                        toplabel.place(x=w / 70, y=h / 2.2)
                        sizeValue=1
                    elif lettuceA>54 and lettuceA<=90:
                        toplabel = Label(self.labelframe, text="2 Week", font=int(w / 102))
                        toplabel.place(x=w / 70, y=h / 2.2)
                        sizeValue=2
                    elif lettuceA>90 and lettuceA<=143:
                        toplabel = Label(self.labelframe, text="3 Week", font=int(w / 102))
                        toplabel.place(x=w / 70, y=h / 2.2)
                        sizeValue=3
                    elif lettuceA>143 and lettuceA<=182:
                        toplabel = Label(self.labelframe, text="4 Week", font=int(w / 102))
                        toplabel.place(x=w / 70, y=h / 2.2)
                        sizeValue=4
                    elif lettuceA > 182 and lettuceA <= 270:
                        toplabel = Label(self.labelframe, text="5 Week", font=int(w / 102))
                        toplabel.place(x=w / 70, y=h / 2.2)
                        sizeValue = 5
                    elif lettuceA>270 and lettuceA<=399:
                        toplabel = Label(self.labelframe, text="6 Week", font=int(w / 102))
                        toplabel.place(x=w / 70, y=h / 2.2)
                        sizeValue = 6
                    elif lettuceA>399 and lettuceA<=525:
                        toplabel = Label(self.labelframe, text="7 Week", font=int(w / 102))
                        toplabel.place(x=w / 70, y=h / 2.2)
                        sizeValue = 7
                    else:
                        toplabel = Label(self.labelframe, text="Uknown", font=int(w / 102))
                        toplabel.place(x=w / 70, y=h / 5)
                        sizeValue = 0
                except FileNotFoundError:
                        toplabel = Label(self.labelframe, text="LOADING... (Press Again Later!)", font=int(w / 102))
                        toplabel.place(x=w / 3.5, y=h / 2.5)

                weeksLeft = 7 - sizeValue
                toplabel = Label(self.labelframe, text= str(weeksLeft) + " Weeks Left", font=int(w / 102))
                toplabel.place(x=w / 70, y=h / 2)

                currentMonth = text = f"{dt.datetime.now():%b}"
                currentYear = int(f"{dt.datetime.now():%Y}")
                try:
                    ePrice_data = pd.read_excel(r'excelFiles/elektrik.xls')
                    all_data = pd.DataFrame(ePrice_data)
                    all_data['Condition1'] = all_data['Month'].apply(lambda x: 'True' if x == currentMonth else 'False')
                    all_data['Condition2'] = all_data['Year'].apply(lambda x: 'True' if x == currentYear else 'False')
                    all_data['Condition3'] = np.where((all_data['Condition1'] == all_data['Condition2']),all_data['Condition1'], "False")
                    showPrice = all_data.loc[all_data['Condition3'] == 'True']['Price'].values[0]

                    eUsedData= pd.read_excel(r'excelFiles/electricityUsed.xls')
                    all_data = pd.DataFrame(eUsedData)
                    all_data['Condition1'] = all_data['Month'].apply(lambda x: 'True' if x == currentMonth else 'False')
                    all_data['Condition2'] = all_data['Year'].apply(lambda x: 'True' if x == currentYear else 'False')
                    all_data['Condition3'] = np.where((all_data['Condition1'] == all_data['Condition2']),all_data['Condition1'], "False")
                    electricityUsed = all_data.loc[all_data['Condition3'] == 'True']['Used'].values[0]

                    estimatedData= pd.read_excel(r'excelFiles/electricityPerWeek.xls')
                    all_data = pd.DataFrame(estimatedData)
                    all_data['Condition1'] = all_data['Month'].apply(lambda x: 'True' if x == currentMonth else 'False')
                    all_data['Condition2'] = all_data['Year'].apply(lambda x: 'True' if x == currentYear else 'False')
                    all_data['Condition3'] = np.where((all_data['Condition1'] == all_data['Condition2']),all_data['Condition1'], "False")
                    estimatedWeek = all_data.loc[all_data['Condition3'] == 'True']['Used'].values[0]

                    #print (showPrice)
                    #print(electricityUsed)

                    toplabel = Label(self.labelframe, text="Normal Conditions", font=int(w / 102))
                    toplabel.place(x=w / 2.56, y=h / 2.2)

                    toplabel = Label(self.labelframe, text=str(sizeValue) + " Weeks used: " + str(electricityUsed) + "kWh\n" + "Total Price = " + str(electricityUsed*showPrice), font=int(w / 102))
                    toplabel.place(x=w / 2.5, y=h / 2.1)

                    toplabel = Label(self.labelframe, text=str(weeksLeft) + " Weeks Left " + "\nEstimated kWh per week: ~" + str(estimatedWeek) +"\nEstimated kWh: ~" + str(weeksLeft*estimatedWeek) + "\nEstimated price: ~" + str(weeksLeft*estimatedWeek*showPrice), font=int(w / 102))
                    toplabel.place(x=w / 2.5, y=h / 1.9)

                    estimatedTotalPrice= (electricityUsed*showPrice) + (weeksLeft*estimatedWeek+showPrice)

                    toplabel = Label(self.labelframe, text="Estimated total price of electricity for harvest: ~" + str(estimatedTotalPrice), font=int(w / 102))
                    toplabel.place(x=w / 2.6, y=h / 1.7)

                except FileNotFoundError:
                    pass

                try:
                    lettucePrice = pd.read_excel(r'excelFiles/lettuce.xls')
                    all_data = pd.DataFrame(lettucePrice)
                    all_data['Condition1'] = all_data['Month'].apply(lambda x: 'True' if x == currentMonth else 'False')
                    all_data['Condition2'] = all_data['Year'].apply(lambda x: 'True' if x == currentYear else 'False')
                    all_data['Condition3'] = np.where((all_data['Condition1'] == all_data['Condition2']),all_data['Condition1'], "False")
                    monthLettucePrice = all_data.loc[all_data['Condition3'] == 'True']['Price'].values[0]
                    print(monthLettucePrice)
                except FileNotFoundError:
                    pass



            elif plant_id==3:
                self.labelframe = LabelFrame(self.statusWindow, text="Pepper", width=int(w / 1.2), height=int(h))
                self.labelframe.place(x=w / 4.6, y=h / 25)




        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        w = int(screen_width)
        h = int(screen_height)
        self.statusWindow = Toplevel()  # Toplevel
        self.statusWindow.geometry(f'{int(w)}x{int(h)}')
        self.statusWindow.configure(bg="#049fd9")
        self.statusWindow.title("Crop Status")




        self.buttonOda = Button(self.statusWindow, text="Tomato", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102),command=lambda: show_status(1))
        self.buttonOda.place(x=w / 55, y=h / 25)

        self.buttonOda = Button(self.statusWindow, text="Lettuce", fg="blue", bg="#fceea7", width=int(w / 102.4), height=int(h / 288), font=int(w / 102),command=lambda: show_status(2))
        self.buttonOda.place(x=w / 55, y=h / 8)

        self.buttonOda = Button(self.statusWindow, text="Pepper", fg="blue", bg="#fceea7", width=int(w / 102.4), height=int(h / 288), font=int(w / 102),command=lambda: show_status(3))
        self.buttonOda.place(x=w / 55, y=h / 4.8)

        self.close_button = Button(self.statusWindow, text="Close", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102), command=self.statusWindow.destroy)
        self.close_button.place(x=w / 50, y=h / 1.4)

################################################
# Green House controls button
# Will send commands to labview
################################################
    def greenHouseControls(self):
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        w = int(screen_width)
        h = int(screen_height)
        self.controlWindow = Toplevel()  # Toplevel
        self.controlWindow.geometry(f'{int(w)}x{int(h)}')
        self.controlWindow.configure(bg="#049fd9")
        self.controlWindow.title("Greenhouse Controls")

        self.close_button = Button(self.controlWindow, text="Close", fg="blue", bg="#fceea7", width=int(w / 102.4), height=int(h / 288), font=int(w / 102), command=self.controlWindow.destroy)
        self.close_button.place(x=w / 50, y=h / 1.4)

################################################
#Warning Button -> Data Check
#Checking if all data files retrieved
################################################
    def greenHouseWarnings(self):

        def dataCheck():

            #Checking if Price data is available
            try:
                ePrice_data = pd.read_excel(r'excelFiles/elektrik.xls')
                ePrice=1
            except FileNotFoundError:
                ePrice=0
                datetime_object = str(dt.datetime.now())
                strIssue="Electricity price data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path='excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False,header=False)
                    writer.save()
                except FileNotFoundError:
                    pass



            try:
                lettucePrice = pd.read_excel(r'excelFiles/lettuce.xls')
                lPrice=1
            except FileNotFoundError:
                lPrice=0
                datetime_object = str(dt.datetime.now())
                strIssue="Lettuce price data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path='excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False,header=False)
                    writer.save()
                except FileNotFoundError:
                    pass

            try:
                tomatoPrice = pd.read_excel(r'excelFiles/tomato.xls')
                tPrice=1
            except FileNotFoundError:
                tPrice=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Tomato price data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass

            try:
                pepperPrice = pd.read_excel(r'excelFiles/pepper.xls')
                pPrice=1
            except FileNotFoundError:
                pPrice=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Pepper price data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass

            try:
                result = docx2txt.process("instructions/cleaning.docx")
                instruction=1
            except FileNotFoundError:
                instruction=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Instruction data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass

            # Checking if Cabin data is available
            try:
                self.pic1 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 1_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 1_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                room1=1
            except FileNotFoundError:
                room1=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Room1 data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            try:
                self.pic1 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 2_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 2_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                room2=1
            except FileNotFoundError:
                room2=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Room2 data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            try:
                self.pic1 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 3_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 3_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                room3=1
            except FileNotFoundError:
                room3=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Room3 data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            try:
                self.pic1 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 4_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 4_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                room4=1
            except FileNotFoundError:
                room4=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Room4 data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            try:
                self.pic1 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 5_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 5_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                room5=1
            except FileNotFoundError:
                room5=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Room5 data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            try:
                self.pic1 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 6_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 6_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                room6=1
            except FileNotFoundError:
                room6=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Room6 data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            try:
                self.pic1 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 7_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 7_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                room7=1
            except FileNotFoundError:
                room7=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Room7 data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            try:
                self.pic1 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 8_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 8_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                room8=1
            except FileNotFoundError:
                room8=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Room8 data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            try:
                self.pic1 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 9_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 9_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                room9=1
            except FileNotFoundError:
                room9=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Room9 data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            try:
                self.pic1 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 10_0.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                self.pic2 = ImageTk.PhotoImage( Image.open("PSO_Check/Oda 10_1.png").resize((int(w / 2.9), int(h / 2.9)), Image.ANTIALIAS))
                room10=1
            except FileNotFoundError:
                room10=0
                datetime_object = str(dt.datetime.now())
                strIssue = "Room10 data not found!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass



            self.labelframe = LabelFrame(self.warningWindow, text="Data Check", width=int(w / 1.2), height=int(h))
            self.labelframe.place(x=w / 4.6, y=h / 25)

            if ePrice==1:
                toplabel = Label(self.labelframe, text="Electricity Price------>OK")
                toplabel.place(x=w / 70, y=h / 60)
            else:
                toplabel = Label(self.labelframe, text="Electricity Price------>Not Found")
                toplabel.place(x=w / 70, y=h / 60)

            if lPrice==1:
                toplabel = Label(self.labelframe, text="Lettuce Price------>OK")
                toplabel.place(x=w / 70, y=h / 30)
            else:
                toplabel = Label(self.labelframe, text="Lettuce Price------>Not Found")
                toplabel.place(x=w / 70, y=h / 30)

            if tPrice == 1:
                toplabel = Label(self.labelframe, text="Tomato Price------>OK")
                toplabel.place(x=w / 70, y=h / 20)
            else:
                toplabel = Label(self.labelframe, text="Tomato Price------>Not Found")
                toplabel.place(x=w / 70, y=h / 20)

            if pPrice == 1:
                toplabel = Label(self.labelframe, text="Pepper Price------>OK")
                toplabel.place(x=w / 70, y=h / 15)
            else:
                toplabel = Label(self.labelframe, text="Pepper Price------>Not Found")
                toplabel.place(x=w / 70, y=h / 15)

            if instruction == 1:
                toplabel = Label(self.labelframe, text="Instruction------>OK")
                toplabel.place(x=w / 70, y=h / 12.2)
            else:
                toplabel = Label(self.labelframe, text="Instruction------>Not Found")
                toplabel.place(x=w / 70, y=h / 12.2)

            if room1 == 1:
                toplabel = Label(self.labelframe, text="Cabin 1------>OK")
                toplabel.place(x=w / 70, y=h / 10.5)
            else:
                toplabel = Label(self.labelframe, text="Cabin 1----->Not Found")
                toplabel.place(x=w / 70, y=h / 10.5)

            if room2 == 1:
                toplabel = Label(self.labelframe, text="Cabin 2------>OK")
                toplabel.place(x=w / 70, y=h / 9)
            else:
                toplabel = Label(self.labelframe, text="Cabin 2------>Not Found")
                toplabel.place(x=w / 70, y=h / 9)

            if room3 == 1:
                toplabel = Label(self.labelframe, text="Cabin 3------>OK")
                toplabel.place(x=w / 70, y=h / 8)
            else:
                toplabel = Label(self.labelframe, text="Cabin 3------>Not Found")
                toplabel.place(x=w / 70, y=h / 8)

            if room4 == 1:
                toplabel = Label(self.labelframe, text="Cabin 4------>OK")
                toplabel.place(x=w / 70, y=h / 7.2)
            else:
                toplabel = Label(self.labelframe, text="Cabin 4------>Not Found")
                toplabel.place(x=w / 70, y=h / 7.2)

            if room5 == 1:
                toplabel = Label(self.labelframe, text="Cabin 5------>OK")
                toplabel.place(x=w / 70, y=h / 6.5)
            else:
                toplabel = Label(self.labelframe, text="Cabin 5------>Not Found")
                toplabel.place(x=w / 70, y=h / 6.5)

            if room6 == 1:
                toplabel = Label(self.labelframe, text="Cabin 6------>OK")
                toplabel.place(x=w / 70, y=h / 5.9)
            else:
                toplabel = Label(self.labelframe, text="Cabin 6------>Not Found")
                toplabel.place(x=w / 70, y=h / 5.9)

            if room7 == 1:
                toplabel = Label(self.labelframe, text="Cabin 7------>OK")
                toplabel.place(x=w / 70, y=h / 5.5)
            else:
                toplabel = Label(self.labelframe, text="Cabin 7------>Not Found")
                toplabel.place(x=w / 70, y=h / 5.5)

            if room8 == 1:
                toplabel = Label(self.labelframe, text="Cabin 8------>OK")
                toplabel.place(x=w / 70, y=h / 5.1)
            else:
                toplabel = Label(self.labelframe, text="Cabin 8------>Not Found")
                toplabel.place(x=w / 70, y=h / 5.1)

            if room9 == 1:
                toplabel = Label(self.labelframe, text="Cabin 9------>OK")
                toplabel.place(x=w / 70, y=h / 4.8)
            else:
                toplabel = Label(self.labelframe, text="Cabin 9------>Not Found")
                toplabel.place(x=w / 70, y=h / 4.8)

            if room10 == 1:
                toplabel = Label(self.labelframe, text="Cabin 10------>OK")
                toplabel.place(x=w / 70, y=h / 4.5)
            else:
                toplabel = Label(self.labelframe, text="Cabin 10------>Not Found")
                toplabel.place(x=w / 70, y=h / 4.5)

            toplabel = Label(self.labelframe,text="If Not Found contact administrator for data check")
            toplabel.place(x=w / 5, y=h / 60)
            try:
                self.picOk = ImageTk.PhotoImage( Image.open("pics/ok.png").resize((int(w / 8), int(h / 4)), Image.ANTIALIAS))
                self.picWarn = ImageTk.PhotoImage( Image.open("pics/warning.jpg").resize((int(w / 6), int(h / 8)), Image.ANTIALIAS))
                if ((ePrice and tPrice and lPrice and pPrice and instruction and room1 and room2 and room3 and room4 and room5 and room5 and room6 and room7 and room8 and room9 and room10) == 1):
                    toplabel = Label(self.labelframe, image=self.picOk)
                    toplabel.place(x=w / 5, y=h / 20)
                else:
                    toplabel = Label(self.labelframe, image=self.picWarn)
                    toplabel.place(x=w / 5, y=h / 10)
            except FileNotFoundError:
                pass

################################################
# Warning Button -> Weather conditions
# Connected to API (openweatermap)
# checking live info for weather
################################################

        def weatherConditions():

            self.labelframe = LabelFrame(self.warningWindow, text="Weather Conditions", width=int(w / 1.2), height=int(h))
            self.labelframe.place(x=w / 4.6, y=h / 25)

            APIKEY = 'ed4a6d93e2c3f1bbbcd7b31833885c0b'  # your API Key here as string
            OpenWMap = pyowm.OWM(APIKEY)  # Use API key to get data
            Weather = OpenWMap.weather_at_place('Ankara')  # give where you need to see the weather
            Data = Weather.get_weather()
            temp = Data.get_temperature(unit='celsius')
            humidity = Data.get_humidity()
            wind = Data.get_wind()
            cloud = Data.get_clouds()
            #temp['temp']=80
            #wind['speed']=12

            tempLabel = Label(self.labelframe, text="Current Temperature: " + str(temp['temp']) + chr(176) +"C")
            tempLabel.place(x=w / 70, y=h / 60)

            humLabel = Label(self.labelframe, text="Current Humidity: " + str(humidity) + "%")
            humLabel.place(x=w / 70, y=h / 30)

            windSpeed = Label(self.labelframe, text="Current Wind Speed: " + str(wind['speed']) + " m/s")
            windSpeed.place(x=w / 70, y=h / 20)

            windDeg = Label(self.labelframe, text="Current Wind Degree: " + str(wind['deg']) + chr(176))
            windDeg.place(x=w / 70, y=h / 15)

            cloudLabel = Label(self.labelframe, text="Current Cloud percentage: " + str(cloud) + "%")
            cloudLabel.place(x=w / 70, y=h / 12.2)

            Weatherforecast = OpenWMap.three_hours_forecast('Ankara')
            rain = Weatherforecast.will_have_rain()  # forecast rain
            #rain=False
            sun = Weatherforecast.will_have_sun()  # forecast sun
            cloud = Weatherforecast.will_have_clouds()  # forecast clouds
            snow=Weatherforecast.will_have_snow()
            #snow=True
            fog = Weatherforecast.will_have_fog()
            tornado = Weatherforecast.will_have_tornado()
            hurricane = Weatherforecast.will_have_hurricane()
            storm = Weatherforecast.will_have_storm()
            #storm=True

            rainLabel = Label(self.labelframe, text="There will be rain : " + str(rain))
            rainLabel.place(x=w / 5, y=h / 60)
            sunLabel = Label(self.labelframe, text="There will be sun : " + str(sun))
            sunLabel.place(x=w / 5, y=h / 30)
            cloudLabel = Label(self.labelframe, text="There will be cloud: " + str(cloud))
            cloudLabel.place(x=w / 5, y=h / 20)
            snowLabel = Label(self.labelframe, text="There will be snow: " + str(snow))
            snowLabel.place(x=w / 5, y=h / 15)
            fogLabel = Label(self.labelframe, text="There will be fog: " + str(fog))
            fogLabel.place(x=w / 3, y=h / 60)
            tornadoLabel = Label(self.labelframe, text="There will be tornado: " + str(tornado))
            tornadoLabel.place(x=w / 3, y=h / 30)
            stormLabel = Label(self.labelframe, text="There will be storm: " + str(storm))
            stormLabel.place(x=w / 3, y=h / 20)
            try:
                self.weatherPic = ImageTk.PhotoImage(Image.open("pics/weather.png").resize((int(w / 8), int(h / 4)), Image.ANTIALIAS))
                weatherLabel = Label(self.labelframe, image=self.weatherPic)
                weatherLabel.place(x=w / 1.55, y=h / 98)
            except FileNotFoundError:
                pass

            if str(rain)=="True":
                rainWarnLabel = Label(self.labelframe, text="Rain expected during 5 days period!", font=int(w / 102))
                rainWarnLabel.place(x=w / 30, y=h /2.5)
                try:
                    self.picWarn = ImageTk.PhotoImage(Image.open("pics/warning.jpg").resize((int(w / 6), int(h / 8)), Image.ANTIALIAS))
                    warnlabel = Label(self.labelframe, image=self.picWarn)
                    warnlabel.place(x=w / 30, y=h / 4.5)
                except FileNotFoundError:
                    pass

            if str(snow)=="True":
                snowWarnLabel = Label(self.labelframe, text="Snow expected during 5 days period!",font=int(w / 102))
                snowWarnLabel.place(x=w / 30, y=h /2.4)
                try:
                    self.picWarn = ImageTk.PhotoImage(Image.open("pics/warning.jpg").resize((int(w / 6), int(h / 8)), Image.ANTIALIAS))
                    warnlabel = Label(self.labelframe, image=self.picWarn)
                    warnlabel.place(x=w / 30, y=h / 4.5)
                except FileNotFoundError:
                    pass

            if str(storm)=="True":
                stormWarnLabel = Label(self.labelframe, text="Storm expected during 5 days period!",font=int(w / 102))
                stormWarnLabel.place(x=w / 30, y=h /2.3)
                try:
                    self.picWarn = ImageTk.PhotoImage(Image.open("pics/warning.jpg").resize((int(w / 6), int(h / 8)), Image.ANTIALIAS))
                    warnlabel = Label(self.labelframe, image=self.picWarn)
                    warnlabel.place(x=w / 30, y=h / 4.5)
                except FileNotFoundError:
                    pass
            if str(rain)=="False" and str(snow)=="False" and str(storm)=="False":
                wWarnLabel = Label(self.labelframe, text="No rain/snow/storm expected during 5 days period" ,font=int(w / 102))
                wWarnLabel.place(x=w / 15, y=h / 2.5)
                try:
                    self.picWarn = ImageTk.PhotoImage(Image.open("pics/ok.png").resize((int(w / 10), int(h / 6)), Image.ANTIALIAS))
                    warnlabel = Label(self.labelframe, image=self.picWarn)
                    warnlabel.place(x=w / 15, y=h / 5.1)
                except FileNotFoundError:
                    pass

            if temp['temp']<=-18:
                tempWarnLevel=Label(self.labelframe, text="Low Temperature!!!", font=int(w / 102))
                tempWarnLevel.place(x=w / 3.3, y=h / 2.5)
                try:
                    self.tempPicWarn = ImageTk.PhotoImage(Image.open("pics/freeze.png").resize((int(w / 10), int(h / 6)), Image.ANTIALIAS))
                    tempPiclabel = Label(self.labelframe, image=self.tempPicWarn)
                    tempPiclabel.place(x=w / 3.5, y=h / 5.1)
                except FileNotFoundError:
                    pass
            elif temp['temp']>=35:
                tempWarnLevel=Label(self.labelframe, text="High Temperature!!!", font=int(w / 102))
                tempWarnLevel.place(x=w / 3.3, y=h / 2.5)
                try:
                    self.tempPicWarn = ImageTk.PhotoImage(Image.open("pics/highTemp.png").resize((int(w / 10), int(h / 6)), Image.ANTIALIAS))
                    tempPiclabel = Label(self.labelframe, image=self.tempPicWarn)
                    tempPiclabel.place(x=w / 3.5, y=h / 5.1)
                except FileNotFoundError:
                    pass
            else:
                tempWarnLevel=Label(self.labelframe, text="Standard Temperature", font=int(w / 102))
                tempWarnLevel.place(x=w / 3.3, y=h / 2.5)
                try:
                    self.tempPicWarn = ImageTk.PhotoImage(Image.open("pics/ok.png").resize((int(w / 10), int(h / 6)), Image.ANTIALIAS))
                    tempPiclabel = Label(self.labelframe, image=self.tempPicWarn)
                    tempPiclabel.place(x=w / 3.5, y=h / 5.1)
                except FileNotFoundError:
                    pass

            if wind['speed']>=12:
                windWarnLevel=Label(self.labelframe, text="High Wind Speed!!!", font=int(w / 102))
                windWarnLevel.place(x=w / 2, y=h / 2.5)
                try:
                    self.windPicWarn = ImageTk.PhotoImage(Image.open("pics/highWind.png").resize((int(w / 10), int(h / 6)), Image.ANTIALIAS))
                    tempPiclabel = Label(self.labelframe, image=self.windPicWarn)
                    tempPiclabel.place(x=w / 2.1, y=h / 5.1)
                except FileNotFoundError:
                    pass
            else:
                windWarnLevel = Label(self.labelframe, text="Standard Wind", font=int(w / 102))
                windWarnLevel.place(x=w / 2, y=h / 2.5)
                try:
                    self.windPicWarn = ImageTk.PhotoImage(
                        Image.open("pics/ok.png").resize((int(w / 10), int(h / 6)), Image.ANTIALIAS))
                    tempPiclabel = Label(self.labelframe, image=self.windPicWarn)
                    tempPiclabel.place(x=w / 2.1, y=h / 5.1)
                except FileNotFoundError:
                    pass

            try:
                self.tempChart = ImageTk.PhotoImage(Image.open("pics/tempChart.png").resize((int(w / 3), int(h / 3)), Image.ANTIALIAS))
                tempPiclabel = Label(self.labelframe, image=self.tempChart)
                tempPiclabel.place(x=w / 5, y=h / 2)
            except FileNotFoundError:
                pass

################################################
# Warning Button -> Internet Check
# Checking if device connected to the internet
################################################
        def internetCheck():
            def connect(host='http://google.com'):
                try:
                    urllib.request.urlopen(host)  # Python 3.x
                    return True
                except:
                    return False
            # test
            if (connect()):
                messagebox.showinfo("Internet", "Connected")
            else:
                messagebox.showerror("Internet", "No internet connection!")
                datetime_object = str(dt.datetime.now())
                strIssue = "No internet connection!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            self.warningWindow.destroy()

################################################
# Warning Button -> Ip address check
# Checking ipv4 of the device
################################################

        def ipV4Check():
            try:
                IP_Address = socket.gethostbyname(socket.gethostname())
                messagebox.showinfo("IPv4 address", "IP Address of your device is " + IP_Address)
            except:
                messagebox.showerror("IPv4 address", "Uknown")
            self.warningWindow.destroy()

        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        w = int(screen_width)
        h = int(screen_height)

################################################
# Warning Button -> Sensors Check
# Sensors info from Labivew stored in excel file
# To check every room if temp/hum sensors are working
################################################
        def sensorsCheck():
            self.labelframe = LabelFrame(self.warningWindow, text="Sensors Check", width=int(w / 1.2),height=int(h))
            self.labelframe.place(x=w / 4.6, y=h / 25)
            room1= pd.read_excel(r'excelFiles/sensorsCheck.xls',sheet_name='Oda1')
            room2 = pd.read_excel(r'excelFiles/sensorsCheck.xls', sheet_name='Oda2')
            room3 = pd.read_excel(r'excelFiles/sensorsCheck.xls', sheet_name='Oda3')
            room4 = pd.read_excel(r'excelFiles/sensorsCheck.xls', sheet_name='Oda4')
            room5 = pd.read_excel(r'excelFiles/sensorsCheck.xls', sheet_name='Oda5')
            room6 = pd.read_excel(r'excelFiles/sensorsCheck.xls', sheet_name='Oda6')
            room7 = pd.read_excel(r'excelFiles/sensorsCheck.xls', sheet_name='Oda7')
            room8 = pd.read_excel(r'excelFiles/sensorsCheck.xls', sheet_name='Oda8')
            room9 = pd.read_excel(r'excelFiles/sensorsCheck.xls', sheet_name='Oda9')
            room10 = pd.read_excel(r'excelFiles/sensorsCheck.xls', sheet_name='Oda10')

            if int(room1['Temp'])==1:
                room1Temp="Working"
                toplabel = Label(self.labelframe, text="Room 1 Temperature------> " + room1Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 60)
            elif int(room1['Temp'])==0:
                room1Temp="Not Working"
                toplabel = Label(self.labelframe, text="Room 1 Temperature------> " + room1Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 60)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room1 temperature sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room1Temp = "Uknown"
                toplabel = Label(self.labelframe, text="Room 1 Temperature------> " + room1Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 60)

            if int(room2['Temp'])==1:
                room2Temp="Working"
                toplabel = Label(self.labelframe, text="Room 2 Temperature------> " + room2Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 30)
            elif int(room2['Temp'])==0:
                room2Temp="Not Working"
                toplabel = Label(self.labelframe, text="Room 2 Temperature------> " + room2Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 30)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room2 temperature sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass

            else:
                room2Temp = "Uknown"
                toplabel = Label(self.labelframe, text="Room 2 Temperature------> " + room2Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 30)

            if int(room3['Temp'])==1:
                room3Temp="Working"
                toplabel = Label(self.labelframe, text="Room 3 Temperature------> " + room3Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 20)
            elif int(room3['Temp'])==0:
                room3Temp="Not Working"
                toplabel = Label(self.labelframe, text="Room 3 Temperature------> " + room3Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 20)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room3 temperature sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room3Temp = "Uknown"
                toplabel = Label(self.labelframe, text="Room 3 Temperature------> " + room3Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 20)

            if int(room4['Temp']) == 1:
                room4Temp = "Working"
                toplabel = Label(self.labelframe, text="Room 4 Temperature------> " + room4Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 15)

            elif int(room4['Temp']) == 0:
                room4Temp = "Not Working"
                toplabel = Label(self.labelframe, text="Room 4 Temperature------> " + room4Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 15)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room4 temperature sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room4Temp = "Uknown"
                toplabel = Label(self.labelframe, text="Room 4 Temperature------> " + room4Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 15)

            if int(room5['Temp']) == 1:
                room5Temp = "Working"
                toplabel = Label(self.labelframe, text="Room 5 Temperature------> " + room5Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 12)
            elif int(room5['Temp']) == 0:
                room5Temp = "Not Working"
                toplabel = Label(self.labelframe, text="Room 5 Temperature------> " + room5Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 12)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room5 temperature sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room5Temp = "Uknown"
                toplabel = Label(self.labelframe, text="Room 5 Temperature------> " + room5Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 12)

            if int(room6['Temp']) == 1:
                room6Temp = "Working"
                toplabel = Label(self.labelframe, text="Room 6 Temperature------> " + room6Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 10)
            elif int(room6['Temp']) == 0:
                room6Temp = "Not Working"
                toplabel = Label(self.labelframe, text="Room 6 Temperature------> " + room6Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 10)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room6 temperature sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room6Temp = "Uknown"
                toplabel = Label(self.labelframe, text="Room 6 Temperature------> " + room6Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 10)

            if int(room7['Temp']) == 1:
                room7Temp = "Working"
                toplabel = Label(self.labelframe, text="Room 7 Temperature------> " + room7Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 8.5)
            elif int(room7['Temp']) == 0:
                room7Temp = "Not Working"
                toplabel = Label(self.labelframe, text="Room 7 Temperature------> " + room7Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 8.5)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room7 temperature sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room7Temp = "Uknown"
                toplabel = Label(self.labelframe, text="Room 7 Temperature------> " + room7Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 8.5)

            if int(room8['Temp']) == 1:
                room8Temp = "Working"
                toplabel = Label(self.labelframe, text="Room 8 Temperature------> " + room8Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 7.5)
            elif int(room8['Temp']) == 0:
                room8Temp = "Not Working"
                toplabel = Label(self.labelframe, text="Room 8 Temperature------> " + room8Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 7.5)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room8 temperature sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room8Temp = "Uknown"
                toplabel = Label(self.labelframe, text="Room 8 Temperature------> " + room8Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 7.5)

            if int(room9['Temp']) == 1:
                room9Temp = "Working"
                toplabel = Label(self.labelframe, text="Room 9 Temperature------> " + room9Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 6.7)
            elif int(room9['Temp']) == 0:
                room9Temp = "Not Working"
                toplabel = Label(self.labelframe, text="Room 9 Temperature------> " + room9Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 6.7)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room9 temperature sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room9Temp = "Uknown"
                toplabel = Label(self.labelframe, text="Room 9 Temperature------> " + room9Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 6.7)

            if int(room10['Temp']) == 1:
                room10Temp = "Working"
                toplabel = Label(self.labelframe, text="Room 10 Temperature------> " + room10Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 6)
            elif int(room10['Temp']) == 0:
                room10Temp = "Not Working"
                toplabel = Label(self.labelframe, text="Room 10 Temperature------> " + room10Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 6)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room10 temperature sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room10Temp = "Uknown"
                toplabel = Label(self.labelframe, text="Room 10 Temperature------> " + room10Temp, font=int(w / 10))
                toplabel.place(x=w / 70, y=h / 6)

#######################
#FOR HUMIDITY##########
#######################

            if int(room1['Hum']) == 1:
                room1Hum = "Working"
                toplabel = Label(self.labelframe, text="Room 1 Humidity------> " + room1Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 60)
            elif int(room1['Hum']) == 0:
                room1Hum = "Not Working"
                toplabel = Label(self.labelframe, text="Room 1 Humidity------> " + room1Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 60)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room1 humidity sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room1Hum = "Uknown"
                toplabel = Label(self.labelframe, text="Room 1 Humidity------> " + room1Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 60)

            if int(room2['Hum']) == 1:
                room2Hum = "Working"
                toplabel = Label(self.labelframe, text="Room 2 Humidity------> " + room2Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 30)
            elif int(room2['Hum']) == 0:
                room2Hum = "Not Working"
                toplabel = Label(self.labelframe, text="Room 2 Humidity------> " + room2Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 30)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room2 humidity sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room2Hum = "Uknown"
                toplabel = Label(self.labelframe, text="Room 2 Humidity------> " + room2Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 30)

            if int(room3['Hum']) == 1:
                room3Hum = "Working"
                toplabel = Label(self.labelframe, text="Room 3 Humidity------> " + room3Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 20)
            elif int(room3['Hum']) == 0:
                room3Hum = "Not Working"
                toplabel = Label(self.labelframe, text="Room 3 Humidity------> " + room3Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 20)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room3 humidity sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room3Hum = "Uknown"
                toplabel = Label(self.labelframe, text="Room 3 Humidity------> " + room3Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 20)

            if int(room4['Hum']) == 1:
                room4Hum = "Working"
                toplabel = Label(self.labelframe, text="Room 4 Humidity------> " + room4Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 15)
            elif int(room4['Hum']) == 0:
                room4Hum = "Not Working"
                toplabel = Label(self.labelframe, text="Room 4 Humidity------> " + room4Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 15)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room4 humidity sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room4Hum = "Uknown"
                toplabel = Label(self.labelframe, text="Room 4 Humidity------> " + room4Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 15)

            if int(room5['Hum']) == 1:
                room5Hum = "Working"
                toplabel = Label(self.labelframe, text="Room 5 Humidity------> " + room5Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 12)
            elif int(room5['Hum']) == 0:
                room5Hum = "Not Working"
                toplabel = Label(self.labelframe, text="Room 5 Humidity------> " + room5Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 12)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room5 humidity sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room5Hum = "Uknown"
                toplabel = Label(self.labelframe, text="Room 5 Humidity------> " + room5Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 12)

            if int(room6['Hum']) == 1:
                room6Hum = "Working"
                toplabel = Label(self.labelframe, text="Room 6 Humidity------> " + room6Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 10)
            elif int(room6['Hum']) == 0:
                room6Hum = "Not Working"
                toplabel = Label(self.labelframe, text="Room 6 Humidity------> " + room6Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 10)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room6 humidity sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room6Hum = "Uknown"
                toplabel = Label(self.labelframe, text="Room 6 Humidity------> " + room6Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 10)

            if int(room7['Hum']) == 1:
                room7Hum = "Working"
                toplabel = Label(self.labelframe, text="Room 7 Humidity------> " + room7Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 8.5)
            elif int(room7['Hum']) == 0:
                room7Hum = "Not Working"
                toplabel = Label(self.labelframe, text="Room 7 Humidity------> " + room7Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 8.5)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room7 humidity sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room7Hum = "Uknown"
                toplabel = Label(self.labelframe, text="Room 7 Humidity------> " + room7Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 8.5)

            if int(room8['Hum']) == 1:
                room8Hum = "Working"
                toplabel = Label(self.labelframe, text="Room 8 Humidity------> " + room8Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 7.5)
            elif int(room8['Hum']) == 0:
                room8Hum = "Not Working"
                toplabel = Label(self.labelframe, text="Room 8 Humidity------> " + room8Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 7.5)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room8 humidity sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room8Hum = "Uknown"
                toplabel = Label(self.labelframe, text="Room 8 Humidity------> " + room8Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 7.5)


            if int(room9['Hum']) == 1:
                room9Hum = "Working"
                toplabel = Label(self.labelframe, text="Room 9 Humidity------> " + room9Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 6.7)
            elif int(room9['Hum']) == 0:
                room9Hum = "Not Working"
                toplabel = Label(self.labelframe, text="Room 9 Humidity------> " + room9Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 6.7)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room9 humidity sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room9Hum = "Uknown"
                toplabel = Label(self.labelframe, text="Room 9 Humidity------> " + room9Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 6.7)

            if int(room10['Hum']) == 1:
                room10Hum = "Working"
                toplabel = Label(self.labelframe, text="Room 10 Humidity------> " + room10Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 6)
            elif int(room10['Hum']) == 0:
                room10Hum = "Not Working"
                toplabel = Label(self.labelframe, text="Room 10 Humidity------> " + room10Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 6)
                datetime_object = str(dt.datetime.now())
                strIssue = "Room10 humidity sensor is not working!"
                df = pd.DataFrame({'Issue': [strIssue], 'Time': [datetime_object]})
                try:
                    path = 'excelFiles/historyLogs.xlsx'
                    book = load_workbook(path)
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    for sheetname in writer.sheets:
                        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                    index=False, header=False)
                    writer.save()
                except FileNotFoundError:
                    pass
            else:
                room10Hum = "Uknown"
                toplabel = Label(self.labelframe, text="Room 10 Humidity------> " + room10Hum, font=int(w / 10))
                toplabel.place(x=w / 5, y=h / 6)

##################################
#Warning Button->Open History LOGS
##################################


        def openHistoryLogs():
            os.system("start EXCEL.EXE excelFiles/historyLogs.xlsx")

##################################
# Inside Warning Button functions
##################################

        self.warningWindow = Toplevel()  # Toplevel
        self.warningWindow.geometry(f'{int(w)}x{int(h)}')
        self.warningWindow.configure(bg="#049fd9")
        self.warningWindow.title("Warnings")

        self.buttonOda = Button(self.warningWindow, text="Weather Conditions", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102),command=weatherConditions)
        self.buttonOda.place(x=w / 55, y=h / 25)

        self.buttonOda = Button(self.warningWindow, text="Data Check", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102), command=dataCheck)
        self.buttonOda.place(x=w / 55, y=h / 8)

        self.buttonOda = Button(self.warningWindow, text="Internet Check", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102), command= internetCheck)
        self.buttonOda.place(x=w / 55, y=h / 4.8)

        self.buttonOda = Button(self.warningWindow, text="IPv4 address Check", fg="blue", bg="#fceea7", width=int(w / 102.4),height=int(h / 288), font=int(w / 102), command= ipV4Check)
        self.buttonOda.place(x=w / 55, y=h / 3.4)

        self.buttonOda = Button(self.warningWindow, text="Sensors Check", fg="blue", bg="#fceea7", width=int(w / 102.4), height=int(h / 288), font=int(w / 102),command= sensorsCheck)
        self.buttonOda.place(x=w / 55, y=h / 2.6)

        self.buttonOda = Button(self.warningWindow, text="Open History Logs", fg="blue", bg="#fceea7", width=int(w / 102.4), height=int(h / 288), font=int(w / 102),command= openHistoryLogs)
        self.buttonOda.place(x=w / 55, y=h / 2.1)


        self.close_button = Button(self.warningWindow, text="Close", fg="blue", bg="#fceea7", width=int(w / 102.4), height=int(h / 288), font=int(w / 102), command=self.warningWindow.destroy)
        self.close_button.place(x=w / 50, y=h / 1.4)




def main():

    #os.startfile(r"PSOgraph.exe")
    app = Tk()
    menu = Create_Window(app)
    app.mainloop()


if __name__ == "__main__":
    main()