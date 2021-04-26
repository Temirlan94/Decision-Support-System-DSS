from shutil import copyfile
import time
import openpyxl
import subprocess
from openpyxl.chart import LineChart, Reference, Series, ScatterChart
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
import getpass
import win32com.client as win32
from win32com.client import Dispatch

user = getpass.getuser()

subprocess.call(r'net use del')
subprocess.call(r'net use X: "\\LAB_LENOVO1\LABVIEW Ortak Dosyalar" sHArt11 /user:admin', shell=True)
print("Kanal acildi")

copyfile(r'X:\Sicaklik_Nem_LOG.xlsx', 'C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Sicaklik_Nem_LOG.xlsx')
print("Dosya kopyalandi")

time.sleep(2)

wb = openpyxl.load_workbook('C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Sicaklik_Nem_LOG.xlsx')


def Creator(oda):
    # fontTest = Font(typeface='Calibri')
    # cp = CharacterProperties(latin=fontTest, sz=800)
    # chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

    sheet = wb[oda]
    # for row in range(3,sheet.max_row):
    #     # sheet.cell(row, 1).number_format = 'hh:mm'
    #     if type(sheet.cell(row,3).value) == type(None):
    #         break
    row = sheet.max_row

    print('\n', oda)
    print('Son satir:', row)
    print('Son tarih: ', sheet.cell(row, 1).value)

    current = str(sheet.cell(row, 1).value)[8:10]
    print('Gosterilen gun: ', current)

    index = 0

    for i in range(3, row):
        if str(sheet.cell(i, 1).value)[8:10] == (current or current - 1):
            index = index + 1

    # dummy1 = sheet.cell(row-index+1,2).value
    # dummy2 = sheet.cell(row-index+1,3).value

    sheet.cell(row - index - 1, 1).value = 'Zaman'
    sheet.cell(row - index - 1, 2).value = 'İstenen T'
    sheet.cell(row - index - 1, 3).value = 'Gerceklesen T'
    sheet.cell(row - index - 1, 4).value = 'İstenen RH'
    sheet.cell(row - index - 1, 5).value = 'Gerceklesen RH'

    chart1 = LineChart()
    chart1.width = 36
    chart1.height = 20
    chart1.title = oda + ' Sicaklik'
    chart1.x_axis.title = " Zaman "
    # chart1.x_axis.number_format = 'hh:mm'
    chart1.y_axis.title = " Sicaklik"

    # chart1.style = 13
    chart1.y_axis.crossAx = 500
    chart1.x_axis = DateAxis(crossAx=100)

    values = Reference(sheet, min_col=2, max_col=3, min_row=row - index - 1, max_row=row)
    # values2 = Reference(sheet, min_col = 4, max_col = 5, min_row = row-index-1, max_row = row)

    chart1.add_data(values, titles_from_data=True)

    dates = Reference(sheet, min_col=1, min_row=row - index - 1, max_row=row)

    chart1.set_categories(dates)
    chart1.y_axis.majorUnit = 1
    sheet.add_chart(chart1, "E2")

    #

    chart = LineChart()
    chart.width = 36
    chart.height = 20
    chart.title = oda + ' Nem'
    chart.x_axis.title = " Zaman "
    # chart.x_axis.number_format = 'hh:mm'
    chart.y_axis.title = " Nem"

    # chart.style = 13
    chart.y_axis.crossAx = 500
    chart.x_axis = DateAxis(crossAx=100)

    # values = Reference(sheet, min_col = 2, max_col = 3, min_row = row-index-1, max_row = row)
    values2 = Reference(sheet, min_col=4, max_col=5, min_row=row - index - 1, max_row=row)

    chart.add_data(values2, titles_from_data=True)

    dates = Reference(sheet, min_col=1, min_row=row - index - 1, max_row=row)

    chart.set_categories(dates)
    chart.y_axis.majorUnit = 5
    sheet.add_chart(chart, "E2")

    # sheet.cell(row-index-1,2).value = dummy1
    # sheet.cell(row-index-1,3).value = dummy2


Creator('Oda 1')
Creator('Oda 2')
Creator('Oda 3')
Creator('Oda 4')
Creator('Oda 5')
Creator('Oda 6')
Creator('Oda 7')
Creator('Oda 8')
Creator('Oda 9')
Creator('Oda 10')

wb.save('C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Sicaklik_Nem_LOG.xlsx')
wb.close()

xlApp = Dispatch('Excel.Application')
workbook = xlApp.Workbooks.Open('C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Sicaklik_Nem_LOG.xlsx')
xlApp.DisplayAlerts = False


def Converter(oda):
    xlApp.Sheets(oda).Select()
    xlSheet = xlApp.Sheets(oda)

    i = 0
    for chart in xlSheet.ChartObjects():
        print('Hazirlaniyor')

        chart.CopyPicture()
        # Create new temporary sheet
        xlApp.ActiveWorkbook.Sheets.Add(After=xlApp.ActiveWorkbook.Sheets(3)).Name = "temp_sheet" + str(i)
        temp_sheet = xlApp.ActiveSheet

        # Add chart object to new sheet.
        cht = xlApp.ActiveSheet.ChartObjects().Add(0, 0, 800, 600)
        # Paste copied chart into new object
        cht.Chart.Paste()
        # Export image
        cht.Chart.Export('C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/' + str(oda) + '_' + str(i) + '.png')

        temp_sheet.Delete()
        i = i + 1


Converter('Oda 1')
Converter('Oda 2')
Converter('Oda 3')
Converter('Oda 4')
Converter('Oda 5')
Converter('Oda 6')
Converter('Oda 7')
Converter('Oda 8')
Converter('Oda 9')
Converter('Oda 10')

xlApp.ActiveWorkbook.Save()
xlApp.ActiveWorkbook.Close()
xlApp.DisplayAlerts = True

# sıcaklık
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 1_0.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 2_0.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 3_0.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 4_0.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 5_0.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 6_0.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 7_0.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 8_0.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 9_0.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 10_0.png"', shell=False)
time.sleep(2)

# nem
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 1_1.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 2_1.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 3_1.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 4_1.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 5_1.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 6_1.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 7_1.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 8_1.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 9_1.png"', shell=False)
time.sleep(2)
shell_process = subprocess.Popen('"C:/Users/' + user + '/Desktop/DSS_Python/PSO_Check/Oda 10_1.png"', shell=False)
time.sleep(2)